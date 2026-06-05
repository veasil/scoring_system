"""Sync 卡牌 xlsx 到 cards.db 沙盒（本地直写）或服务器（HTTP）.

用法:
  python scripts/sync_xlsx_cards.py                # 本地直写 data/cards.db
  python scripts/sync_xlsx_cards.py --remote       # 调用 API 写到 BACKEND_URL
  python scripts/sync_xlsx_cards.py --remote --url https://api.example.com --key xxx

环境变量:
  BACKEND_URL  默认 http://localhost:8080
  DEV_KEY      默认 sj0127wqt
"""
import openpyxl, sqlite3, json, time, os, sys, argparse

ARDEN_ID = 1
DB = os.path.join(os.path.dirname(__file__), '..', 'data', 'cards.db')
FILES = [
    r'C:/Users/BibbTwigs/Downloads/经济安全卡牌.xlsx',
]
ATTRS = ['安全力', '脑波力', '实感力', '创心力', '沟通力']


def load_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    H = {h: i for i, h in enumerate(headers)}
    cards = {}
    for r in rows[1:]:
        if r[H['卡牌ID']] is None:
            continue
        cid = int(r[H['卡牌ID']])
        c = cards.setdefault(cid, {
            'key': cid,
            'safetyType': r[H['安全类型']],
            'phase': r[H['阶段']],
            'event': r[H['事件描述']],
            'options': {},
        })
        opt = str(r[H['选项']]).strip()
        eff = {a: (int(r[H[a]]) if r[H[a]] is not None else 0) for a in ATTRS}
        c['options'][opt] = {
            'text': r[H['选项文字']] or '',
            'consequence': r[H['后果描述']] or '',
            'attributeEffects': eff,
        }
    # sort options A,B,C...
    for c in cards.values():
        c['options'] = {k: c['options'][k] for k in sorted(c['options'].keys())}
    return list(cards.values())


def write_local(cards):
    conn = sqlite3.connect(DB)
    conn.execute('BEGIN')
    now = int(time.time() * 1000)
    inserted = updated = 0
    for c in cards:
        options_json = json.dumps(c['options'], ensure_ascii=False)
        existing = conn.execute('SELECT id, version FROM cards WHERE key=?', (c['key'],)).fetchone()
        if existing:
            card_id, ver = existing
            new_ver = (ver or 1) + 1
            conn.execute('''UPDATE cards SET safety_type=?, phase=?, event=?, options_json=?,
                            status='active', version=?, updated_at=?, author_id=?, branch='release'
                            WHERE id=?''',
                         (c['safetyType'], c['phase'], c['event'], options_json,
                          new_ver, now, ARDEN_ID, card_id))
            updated += 1
        else:
            cur = conn.execute('''INSERT INTO cards (key, safety_type, phase, event, options_json,
                                  status, version, created_at, updated_at, author_id, branch)
                                  VALUES (?,?,?,?,?,'active',1,?,?,?, 'release')''',
                               (c['key'], c['safetyType'], c['phase'], c['event'], options_json,
                                now, now, ARDEN_ID))
            card_id = cur.lastrowid
            new_ver = 1
            inserted += 1
        conn.execute('''INSERT INTO card_versions (card_id, key, safety_type, event, phase,
                        options_json, status, version, version_label, created_at, author_id,
                        note, branch, promoted_at)
                        VALUES (?,?,?,?,?,?,'active',?,?,?,?,?, 'main', ?)''',
                     (card_id, c['key'], c['safetyType'], c['event'], c['phase'],
                      options_json, new_ver, '沙盒同步', now, ARDEN_ID,
                      'sync from xlsx by arden', now))
        vid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.execute('UPDATE cards SET current_version_id=? WHERE id=?', (vid, card_id))
    conn.commit()
    conn.close()
    print(f'[local] inserted={inserted} updated={updated}')


def write_remote(cards, url, key):
    import urllib.request
    payload = json.dumps({
        'cards': cards,
        'secretKey': key,
        'authorId': ARDEN_ID,
        'status': 'active',
        'versionLabel': '沙盒同步',
    }, ensure_ascii=False).encode('utf-8')
    req = urllib.request.Request(
        url.rstrip('/') + '/api/admin/cards/bulk-upsert',
        data=payload, method='POST',
        headers={'Content-Type': 'application/json; charset=utf-8'},
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode('utf-8')
            print('[remote]', resp.status, body)
    except urllib.error.HTTPError as e:
        print('[remote ERROR]', e.code, e.read().decode('utf-8', errors='replace'))
        sys.exit(1)


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--remote', action='store_true', help='POST 到服务器而非写本地 sqlite')
    p.add_argument('--url', default=os.environ.get('BACKEND_URL', 'http://localhost:8080'))
    p.add_argument('--key', default=os.environ.get('DEV_KEY', 'sj0127wqt'))
    args = p.parse_args()

    all_cards = []
    for f in FILES:
        all_cards.extend(load_xlsx(f))
    print(f'Parsed {len(all_cards)} cards from xlsx')

    if args.remote:
        write_remote(all_cards, args.url, args.key)
    else:
        write_local(all_cards)


if __name__ == '__main__':
    main()
