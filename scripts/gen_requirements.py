from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()
ws = wb.active
ts = datetime.now().strftime('%Y%m%d_%H%M')
ws.title = f"需求看板 {ts}"

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
P0_FILL = PatternFill('solid', fgColor='FCE4EC')
P1_FILL = PatternFill('solid', fgColor='FFF3E0')
P2_FILL = PatternFill('solid', fgColor='E8F5E9')
P3_FILL = PatternFill('solid', fgColor='F3E5F5')
DONE_FILL = PatternFill('solid', fgColor='E0E0E0')
DONE_FONT = Font(name='Arial', size=10, color='9E9E9E')
BODY_FONT = Font(name='Arial', size=10)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
THIN = Border(
    left=Side(style='thin', color='BDBDBD'),
    right=Side(style='thin', color='BDBDBD'),
    top=Side(style='thin', color='BDBDBD'),
    bottom=Side(style='thin', color='BDBDBD'),
)

headers = ['序号', '优先级', '需求名称', '具体需求与细节', '技术方案/顾问建议',
           '需要协助点', '截止日期', '交付物', '状态', '备注']
col_widths = [6, 10, 18, 44, 50, 26, 14, 24, 10, 28]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = THIN
    ws.column_dimensions[get_column_letter(i)].width = w

rows = [
    # ========== P0 紧急 (5.17前) ==========
    [1, 'P0-紧急', '得分率算总分',
     '当前总分=五维属性绝对值求和。\n'
     '需改为: 得分率=实际得分/该维度理论满分,\n'
     '总分=五维得分率平均x100。\n'
     '涉及: server.js game/finish, game-review.js sumAttributes(), admin游戏分析页。',
     '技术路线:\n'
     '1. card_groups表新增max_scores_json字段, 缓存各维度理论满分(=组内所有卡各维度最高选项之和)。\n'
     '2. game/finish计算时: 得分率=实际值/max_scores_json对应值, 总分=均值x100。\n'
     '3. 前端game-review.js雷达图坐标轴改0-100%。\n'
     '4. 历史数据: 写迁移脚本按新公式回算, 建议新增score_details_json字段。',
     '确认: 五维权重是否等权(各20%)? 还是自定义权重?',
     '2026-05-17', '后端+前端代码, 迁移脚本', '待开发',
     '影响所有历史场次展示, 需确认是否回算'],

    [2, 'P0-紧急', '活动管理系统',
     '核心: 基于"活动"开展桌游的完整流程。\n'
     '5.17前目标:\n'
     '1. activities表完善(已有基础CRUD)\n'
     '2. 守望者可在前端新增活动\n'
     '3. 填写活动码开展桌游 -> 自动填写地点+游戏模式\n'
     '4. 活动与场次数据自动关联\n'
     '后续: 活动看板展示、数据自动推送给守望者。',
     '技术路线:\n'
     '1. 活动码机制: activities表加code字段(唯一), 守望者创建活动时自动生成或手动填写。\n'
     '2. 开局流程改造: 输入活动码 -> 查询活动 -> 自动带出location+game_mode -> game/start时写入activity_id。\n'
     '3. 前端: public/index.html开局弹窗新增"活动码"输入框, 输入后自动填充。\n'
     '4. API: GET /api/activities/by-code/:code(公开, 返回活动信息); game_sessions表已有location字段。\n'
     '5. 后续看板+推送作为5.24迭代。',
     '1. 活动码格式规则(纯数字? 字母+数字?)\n'
     '2. 使用手册需产出',
     '2026-05-17', '活动码开局流程, API, 前端改造, 使用手册', '待开发',
     '后端CRUD已有, 重点是活动码->开局流程串联'],

    [3, 'P0-紧急', '复盘报告升级',
     '程艳青提供完整写作思路+视觉设计模板, 开发者负责技术实现。\n'
     '1. 按设计模板重构报告HTML结构\n'
     '2. 图文结合: 嵌入数据图表\n'
     '3. 数据面板: 关键指标可视化\n'
     '4. Logo水印+品牌元素\n'
     '5. 5.17基础版, 5.19精修版',
     '技术路线:\n'
     '1. 等程艳青交付设计模板后, 将其转为HTML/CSS模板。\n'
     '2. 现有LLM 4节点pipeline: 改造Node4的HTML组装, 套用新模板。\n'
     '3. 图表: Chart.js服务端渲染为base64图片嵌入(或用canvas2image)。\n'
     '4. 品牌素材上传OSS, 报告模板引用CDN地址。\n'
     '注意: 开发者不设计, 只按模板实现。',
     '【阻塞】等程艳青交付:\n'
     '1. 写作思路文档\n'
     '2. 视觉设计模板(Figma/PSD/PDF)\n'
     '3. Logo源文件+品牌色值',
     '5.17 / 5.19', '报告模板代码, 示例报告截图', '待设计稿',
     '开发被设计稿阻塞, 可先搭框架'],

    [4, 'P0-紧急', '录音转录技术调研',
     '调研: 多人桌游场景下清晰采集并转录为文字。\n'
     '考虑: 多人同时说话/环境噪音/儿童声音/实时性。',
     '调研方向:\n'
     '1. 硬件方案: 会议麦克风阵列(Jabra Speak 750/讯飞听见M2), 支持说话人分离\n'
     '2. 云端ASR: 讯飞实时转写API/阿里云智能语音/Whisper large-v3\n'
     '3. 端侧方案: 录音笔(讯飞SR502)+离线转写\n'
     '4. 关键指标: 说话人识别准确率/儿童语音适配/成本/接入难度\n'
     '产出: 方案对比表格(方案/成本/精度/部署难度/推荐度)。',
     '明确预算范围和场景(一桌几人/时长/是否需要实时)',
     '2026-05-17', '技术调研报告(含对比表格)', '待调研',
     '纯调研不涉及编码'],

    # ========== 已完成 ==========
    [5, '已完成', '注册页面',
     '独立页面register.html收集守望者信息(邀请码+手机+姓名+密码)。',
     '已完成。register.html -> /api/auth/register-with-invite。',
     '-', '-', 'register.html', '已完成', '-'],

    [6, '已完成', '场次数据管理',
     '删除或标记不重要的测试数据, 更精准管理场次。',
     '已完成。admin数据审计页支持逐场次审查。',
     '-', '-', '-', '已完成', '-'],

    # ========== P1 中优先级 ==========
    [7, 'P1-中', '新卡牌批量导入',
     '新版卡牌未做实体, 需提前准备导入工具。\n'
     '支持Excel/CSV -> cards表 + 自动建card_versions。',
     '技术路线:\n'
     '1. admin页面新增"批量导入"tab, 上传Excel文件。\n'
     '2. 后端: 解析Excel -> 校验字段 -> 写入cards+card_versions(v1)。\n'
     '3. 复用三表流程: 导入后status=pending, 需手动promote+release。\n'
     '4. 参照现有卡牌数据.xlsx定义导入模板。\n'
     '5. 先搜索现有Excel解析库(xlsx/exceljs)选型。',
     '确认新卡牌Excel模板格式',
     '2026-05-24', '导入工具(admin页面+API), Excel模板', '待开发',
     '新版卡牌实体未定, 工具先行'],

    [8, 'P1-中', 'TTS语音生成(自动阅读题目)',
     '集成到admin后台的功能:\n'
     '1. 一键给每个卡牌组生成全部卡牌语音\n'
     '2. 用大模型TTS API生成高质量语音\n'
     '3. admin界面可选择音色、语速\n'
     '4. 支持调试: 单卡试听样例\n'
     '5. 生成后上传OSS, 写入cards.audio_url',
     '技术路线:\n'
     '1. TTS API选型(先搜索调研):\n'
     '   - 阿里云智能语音(已有阿里云OSS, 生态兼容)\n'
     '   - OpenAI TTS API(高质量, 支持多音色)\n'
     '   - 讯飞TTS / 火山引擎TTS\n'
     '2. admin新增"语音生成"页面: 选卡牌组->选音色/语速->预览->批量生成。\n'
     '3. 生成流程: 读cards.event文本 -> TTS API -> 上传OSS -> 更新audio_url。\n'
     '4. cards表audio_url字段已预留, 前端播放逻辑待加。',
     '1. TTS API预算和选型决策\n2. 确认朗读内容(仅题目event? 还是含选项?)',
     '2026-05-31', 'admin语音生成页, TTS集成, 批量生成脚本', '待开发',
     'audio_url字段已预留; 先调研TTS API再开发'],

    # ========== P2 低优先级 ==========
    [9, 'P2-低', '数据同步到飞书',
     '游戏场次数据/卡牌数据上传到飞书多维表格, 方便非技术人员查看。',
     '技术路线:\n'
     '1. cards-db.js已预留CARDS_SOURCE=feishu切换机制, 可复用。\n'
     '2. 飞书开放平台"多维表格API"做单向同步(系统->飞书)。\n'
     '3. 实现: admin手动"同步到飞书"按钮 + 可选定时任务。\n'
     '4. 需申请飞书应用(app_id/app_secret)。\n'
     '5. 先搜索飞书多维表格SDK/示例代码。',
     '1. 飞书应用权限审批(提前启动)\n2. 确认同步哪些字段和频率',
     '2026-06-07', '飞书同步模块, 配置文档', '待规划',
     '依赖飞书应用审批, 建议提前申请'],

    [10, 'P2-低', 'Admin脱离Streamlit',
     '当前管理后台Streamlit(Python), 长期迁移到独立前端框架。',
     '技术路线:\n'
     '1. 最大重构项, 不建议一次性迁移。\n'
     '2. enterprise-panel(Vue3+ElementPlus)已验证, 可复用技术栈。\n'
     '3. 渐进策略: 新功能直接做Vue admin, 旧功能按使用频率逐步迁移。\n'
     '4. 迁移优先级: 卡牌管理(最复杂) -> 用户管理 -> 活动管理 -> 其余。\n'
     '5. Streamlit和Vue admin可并行运行, 不停服。',
     '确认是否复用enterprise-panel技术栈',
     '长期(Q3)', '迁移方案文档 -> 逐模块交付', '待规划',
     '建议渐进式, 新功能(如TTS页面)直接做Vue'],

    # ========== P3 远期 ==========
    [11, 'P3-远期', '引导+计分系统合并',
     '现状: 小程序有纯前端引导系统(按步弹出引导流程), 计分系统在web端。\n'
     '目标: 一个电脑屏幕上, 桌游开启后同时计分+引导(可能悬浮窗形式)。\n'
     '合并为一体化桌游主持界面。',
     '技术路线:\n'
     '1. 本质是"桌游主持人工作台": 一个SPA整合引导步骤+实时计分+卡牌朗读。\n'
     '2. 技术选型: Vue3或React SPA, 左侧引导流程+右侧计分面板, 可选悬浮窗模式。\n'
     '3. 依赖: #8 TTS语音生成完成后, 可在引导步骤中自动播放。\n'
     '4. 需先从小程序团队拿到引导流程的完整步骤定义。\n'
     '5. 先搜索类似产品(桌游辅助工具)参考交互设计。',
     '1. 小程序引导流程步骤定义文档\n2. 产品侧确认合并后交互原型',
     '待定', '产品原型 -> 技术方案 -> 开发', '待定义',
     '依赖#8完成; 需拿到小程序引导流程定义'],
]

priority_fills = {
    'P0-紧急': P0_FILL, 'P1-中': P1_FILL, 'P2-低': P2_FILL,
    'P3-远期': P3_FILL, '已完成': DONE_FILL
}

for ri, row in enumerate(rows, 2):
    pfill = priority_fills.get(row[1])
    for ci, v in enumerate(row, 1):
        cell = ws.cell(row=ri, column=ci, value=v)
        cell.font = DONE_FONT if row[1] == '已完成' else BODY_FONT
        cell.alignment = CENTER if ci in (1, 2, 7, 9) else WRAP
        cell.border = THIN
        if pfill and ci == 2:
            cell.fill = pfill
        if row[1] == '已完成':
            cell.fill = DONE_FILL

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:J{len(rows)+1}"
# Row height for content rows
for ri in range(2, len(rows) + 2):
    ws.row_dimensions[ri].height = 120

# --- Sheet 2: 工作流程 ---
ws2 = wb.create_sheet("工作流程")
flow = [
    ['阶段', '时间', '角色', '动作', '产出'],
    ['需求录入', '随时', '产品/运营', '飞书文档记录需求, 同步到本表', '需求行(含优先级+截止日期)'],
    ['周末开发', '周六-周日', '开发', '按优先级开发, 截屏交付; 复杂需求备注操作手册链接', '代码+截图+手册'],
    ['周一检查', '周一', '检查方', '查看截图+操作手册, 反馈问题到飞书', '验收意见/改进点'],
    ['周二对接', '周二', '全员', '技术对接会, 确认验收+排下周计划', '下周排期+更新本表'],
]
for ri, row in enumerate(flow, 1):
    for ci, v in enumerate(row, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        if ri == 1:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        else:
            cell.font = BODY_FONT
        cell.alignment = CENTER
        cell.border = THIN
for ci in range(1, 6):
    ws2.column_dimensions[get_column_letter(ci)].width = 30

# --- Sheet 3: 本周冲刺 ---
ws3 = wb.create_sheet("本周冲刺 5.14-5.17")
sprint = [
    ['日期', '上午', '下午', '晚上', '产出确认'],
    ['5.14(三)今天',
     '需求对齐+表格定稿(本表)',
     '得分率算总分: 后端逻辑(server.js + card_groups)',
     '得分率: 前端改造(game-review.js雷达图)',
     '得分率功能可测试'],
    ['5.15(四)',
     '活动管理: activities表加code字段, API开发',
     '活动管理: 前端开局流程(活动码输入->自动填充)',
     '复盘报告: 确认设计稿状态, 搭HTML框架',
     '活动码开局流程可测试'],
    ['5.16(五)',
     '复盘报告: 按设计稿实现(若收到); 否则做图表嵌入',
     '录音转录: 技术调研+写对比报告',
     '使用手册撰写(活动管理)',
     '调研报告+手册完成'],
    ['5.17(六)',
     '联调测试: 得分率+活动码全流程',
     '截图交付+修bug',
     '提交+自检',
     '全部P0交付'],
]
SPRINT_HEADER = PatternFill('solid', fgColor='2E7D32')
TODAY_FILL = PatternFill('solid', fgColor='E8F5E9')
for ri, row in enumerate(sprint, 1):
    for ci, v in enumerate(row, 1):
        cell = ws3.cell(row=ri, column=ci, value=v)
        if ri == 1:
            cell.font = HEADER_FONT
            cell.fill = SPRINT_HEADER
        else:
            cell.font = BODY_FONT
            if ri == 2:
                cell.fill = TODAY_FILL
        cell.alignment = CENTER
        cell.border = THIN
    if ri > 1:
        ws3.row_dimensions[ri].height = 60
for ci in range(1, 6):
    ws3.column_dimensions[get_column_letter(ci)].width = 32

# --- Sheet 4: 依赖关系 ---
ws4 = wb.create_sheet("依赖与风险")
deps = [
    ['需求', '前置依赖', '阻塞风险', '缓解措施'],
    ['#1 得分率', '无(可独立开发)', '历史数据回算可能影响已展示的成绩', '新增字段, 不改原字段; 回算可延后'],
    ['#2 活动管理', '无(后端API已有)', '活动码与现有开局流程的兼容', '活动码可选填, 不填则走旧流程'],
    ['#3 复盘报告', '【阻塞】程艳青设计稿', '设计稿未到则无法按模板开发', '先搭框架+图表, 设计稿到后套模板'],
    ['#8 TTS语音', '#7卡牌导入(有新卡才有意义)', 'TTS API选型和成本', '先调研API对比, 小批量试用'],
    ['#9 飞书同步', '飞书应用审批(外部依赖)', '审批周期不可控', '本周提交申请, 审批期间做其他需求'],
    ['#10 Admin迁移', '无(但工作量最大)', '迁移期间两套系统并行维护成本', '渐进式: 新功能做Vue, 旧功能按需迁移'],
    ['#11 引导合并', '#8 TTS完成 + 小程序引导定义', '产品形态未定义', '先拿到引导步骤文档再评估'],
]
for ri, row in enumerate(deps, 1):
    for ci, v in enumerate(row, 1):
        cell = ws4.cell(row=ri, column=ci, value=v)
        if ri == 1:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        else:
            cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN
for ci in range(1, 5):
    ws4.column_dimensions[get_column_letter(ci)].width = 36

out = rf'c:\Users\BibbTwigs\OneDrive\Desktop\intern\wqt-auth-backend\WQT需求看板_{ts}.xlsx'
wb.save(out)
print(f"Done: {out}")
